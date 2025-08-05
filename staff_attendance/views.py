from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.conf import settings
import json
import datetime
from django.db.models import Count, Q
import os
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.workbook.protection import WorkbookProtection
from django.http import HttpResponse
from io import BytesIO
import datetime

from .models import City, StaffMember, Attendance, CongeReservation, Zone, UserProfile

# Generate PDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from io import BytesIO
import calendar
import locale

MOIS_FRANCAIS = {
    1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
    5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
    9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
}

# Dictionnaire des jours en français
JOURS_FRANCAIS = {
    'Monday': 'Lundi', 'Tuesday': 'Mardi', 'Wednesday': 'Mercredi',
    'Thursday': 'Jeudi', 'Friday': 'Vendredi', 'Saturday': 'Samedi', 'Sunday': 'Dimanche'
}
def login_view(request):
    """Handle user login"""
    # If user is already authenticated, redirect to attendance page
    if request.user.is_authenticated:
        return redirect('staff_attendance:attendance_page')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if not username or not password:
            return render(request, 'staff_attendance/login.html', {
                'error': 'Veuillez remplir tous les champs'
            })

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            # Create user profile if it doesn't exist
            UserProfile.objects.get_or_create(user=user)
            return redirect('staff_attendance:attendance_page')
        else:
            return render(request, 'staff_attendance/login.html', {
                'error': 'Nom d\'utilisateur ou mot de passe incorrect'
            })

    return render(request, 'staff_attendance/login.html')


@csrf_exempt
def login_api(request):
    """API endpoint for login"""
    if request.method == 'POST':
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            profile = UserProfile.objects.get_or_create(user=user)[0]
            return JsonResponse({
                'success': True,
                'message': f'Bienvenue {user.first_name or user.username}',
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'role': profile.role,
                    'is_admin': profile.is_admin,
                    'can_modify_all': profile.can_modify_all
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Nom d\'utilisateur ou mot de passe incorrect'
            }, status=401)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


def logout_view(request):
    """Handle user logout"""
    logout(request)
    return redirect('staff_attendance:login')



@login_required
def attendance_page(request):
    """Render the main attendance page"""
    cities = City.objects.all()
    profile = UserProfile.objects.get_or_create(user=request.user)[0]

    return render(request, 'staff_attendance/attendance.html', {
        'cities': cities,
        'user_profile': profile,
    })


@login_required
def attendance_history(request):
    """Render the attendance history page"""
    cities = City.objects.all()
    profile = UserProfile.objects.get_or_create(user=request.user)[0]

    return render(request, 'staff_attendance/attendance_history.html', {
        'cities': cities,
        'user_profile': profile,
    })


@login_required
def user_management(request):
    """Render the user management page (admin only)"""
    profile = UserProfile.objects.get_or_create(user=request.user)[0]
    if not profile.is_admin:
        return redirect('staff_attendance:attendance_page')

    users = User.objects.all().select_related('profile')
    zones = Zone.objects.all()

    return render(request, 'staff_attendance/user_management.html', {
        'users': users,
        'zones': zones,
    })

def staff_management(request):
    """Render the staff management page for drag and drop functionality"""
    cities = City.objects.all()
    return render(request, 'staff_attendance/staff_management.html', {
        'cities': cities,
    })


@csrf_exempt
def add_city(request):
    """API endpoint to add a new city to a zone"""
    if request.method == 'POST':
        data = json.loads(request.body)
        city_name = data.get('name', '').strip()
        zone_id = data.get('zone_id')

        if not city_name:
            return JsonResponse({'success': False, 'message': 'Le nom du chantier est requis'}, status=400)

        if not zone_id:
            return JsonResponse({'success': False, 'message': 'La zone est requise'}, status=400)

        # Check if city already exists
        if City.objects.filter(name__iexact=city_name).exists():
            return JsonResponse({'success': False, 'message': 'Ce chantier existe déjà'}, status=400)

        try:
            zone = Zone.objects.get(id=zone_id)
            city = City.objects.create(name=city_name, zone=zone)
            return JsonResponse({
                'success': True,
                'message': f'Chantier "{city_name}" ajouté à la zone "{zone.name}" avec succès',
                'city': {
                    'id': city.id,
                    'name': city.name,
                    'zone_id': zone.id,
                    'zone_name': zone.name,
                    'staff_count': 0
                }
            })
        except Zone.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Zone non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


# Ajoutez ces vues à votre staff_attendance/views.py

@login_required
def department_view(request):
    """Render the department-based attendance view"""
    cities = City.objects.all()
    profile = UserProfile.objects.get_or_create(user=request.user)[0]

    return render(request, 'staff_attendance/department_view.html', {
        'cities': cities,
        'user_profile': profile,
    })


@csrf_exempt
@login_required
def get_department_staff(request):
    """API endpoint to get staff organized by departments"""
    try:
        # Get all staff with attendance data
        today = timezone.now().date()
        staff_members = StaffMember.objects.select_related('city').all()
        attendances = Attendance.objects.filter(date=today).select_related('created_by', 'updated_by')

        # Create attendance lookup
        attendance_dict = {att.staff_member_id: att for att in attendances}

        # Check for active congé reservations
        active_conge = CongeReservation.objects.filter(
            start_date__lte=today,
            end_date__gte=today
        ).values_list('staff_member_id', flat=True)

        # Department mapping - MODIFIEZ CES AFFECTATIONS SELON VOS BESOINS
        DEPARTMENT_MAPPING = {
            'contact': ['160 - DROUOT_Réhabilitation de 821 logts'],  # Remplacez par vos chantiers Contact
            'maintenance': ['001 - Bureau'],  # Remplacez par vos chantiers Maintenance
            'moselle-est': ['015 - Atelier']  # Remplacez par vos chantiers Moselle-Est
        }

        # Organize staff by departments
        departments = {
            'contact': [],
            'maintenance': [],
            'moselle-est': []
        }

        # Process each staff member
        for staff in staff_members:
            attendance = attendance_dict.get(staff.id)

            # Determine status
            if staff.id in active_conge:
                status = 'conge'
            elif attendance:
                if attendance.absence_reason == 'CONGE_STATUS':
                    status = 'conge'
                elif attendance.present is True:
                    status = 'present'
                elif attendance.present is False:
                    status = 'absent'
                else:
                    status = 'undefined'
            else:
                status = 'undefined'

            # Get user info for who made the last change
            last_modified_by = None
            if attendance:
                if attendance.updated_by:
                    last_modified_by = f"{attendance.updated_by.first_name or attendance.updated_by.username}"
                elif attendance.created_by:
                    last_modified_by = f"{attendance.created_by.first_name or attendance.created_by.username}"

            # Create staff data
            staff_data = {
                'id': staff.id,
                'name': staff.name,
                'city': staff.city.name,
                'status': status,
                'absence_reason': attendance.absence_reason if attendance and attendance.present is False and attendance.absence_reason != 'CONGE_STATUS' else None,
                'timestamp': attendance.timestamp.isoformat() if attendance and attendance.timestamp else None,
                'hours_worked': float(attendance.hours_worked) if attendance and attendance.hours_worked else None,
                'grand_deplacement': attendance.grand_deplacement if attendance else False,
                'last_modified_by': last_modified_by
            }

            # Assign to department based on city
            assigned = False
            for dept, cities in DEPARTMENT_MAPPING.items():
                if staff.city.name in cities:
                    departments[dept].append(staff_data)
                    assigned = True
                    break

            # Si un chantier n'est pas assigné à un département, l'ajouter à "non-assigné" ou contact par défaut
            if not assigned:
                # Par défaut, on met dans contact s'il n'est pas assigné
                departments['contact'].append(staff_data)

        return JsonResponse({
            'success': True,
            'departments': departments,
            'department_mapping': DEPARTMENT_MAPPING,
            'date': today.isoformat()
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required
def mark_department_present(request):
    """API endpoint to mark entire department as present"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            department = data.get('department')

            if not department:
                return JsonResponse({'success': False, 'message': 'Département requis'}, status=400)

            # Department mapping (same as above)
            DEPARTMENT_MAPPING = {
                'contact': ['160 - DROUOT_Réhabilitation de 821 logts'],  # Remplacez par vos chantiers Contact
                'maintenance': ['001 - Bureau'],  # Remplacez par vos chantiers Maintenance
                'moselle-est': ['015 - Atelier']
            }

            if department not in DEPARTMENT_MAPPING:
                return JsonResponse({'success': False, 'message': 'Département invalide'}, status=400)

            # Get all cities for this department
            department_cities = DEPARTMENT_MAPPING[department]

            # Get all staff in these cities
            staff_members = StaffMember.objects.filter(city__name__in=department_cities)

            today = timezone.now().date()
            current_time = timezone.now()
            success_count = 0

            # Mark all staff as present
            for staff in staff_members:
                attendance, created = Attendance.objects.get_or_create(
                    staff_member=staff,
                    date=today,
                    defaults={
                        'present': True,
                        'timestamp': current_time,
                        'hours_worked': 8.0,
                        'grand_deplacement': False,
                        'created_by': request.user
                    }
                )

                if not created:
                    attendance.present = True
                    attendance.absence_reason = None
                    attendance.timestamp = current_time
                    attendance.hours_worked = 8.0
                    attendance.grand_deplacement = False
                    attendance.updated_by = request.user
                    attendance.save()

                success_count += 1

            return JsonResponse({
                'success': True,
                'message': f'{success_count} membres du département {department.upper()} marqués présents'
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def update_department_mapping(request):
    """API endpoint to update department-city mapping (admin only)"""
    profile = UserProfile.objects.get_or_create(user=request.user)[0]
    if not profile.is_admin:
        return JsonResponse({'success': False, 'message': 'Accès refusé'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_mapping = data.get('mapping', {})

            # Validate the mapping structure
            required_departments = ['contact', 'maintenance', 'moselle-est']
            for dept in required_departments:
                if dept not in new_mapping:
                    return JsonResponse({
                        'success': False,
                        'message': f'Département {dept} manquant dans la configuration'
                    }, status=400)

            # Validate that all cities exist
            all_cities = []
            for dept, cities in new_mapping.items():
                all_cities.extend(cities)

            existing_cities = City.objects.filter(name__in=all_cities).values_list('name', flat=True)
            missing_cities = set(all_cities) - set(existing_cities)

            if missing_cities:
                return JsonResponse({
                    'success': False,
                    'message': f'Chantiers inexistants: {", ".join(missing_cities)}'
                }, status=400)

            # Pour une version simple, on pourrait stocker cela dans les settings ou une table de configuration
            # Pour maintenant, on retourne simplement le succès
            return JsonResponse({
                'success': True,
                'message': 'Configuration des départements mise à jour avec succès',
                'mapping': new_mapping
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'JSON invalide'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


# Add this to staff_attendance/views.py

def fixed_teams_page(request):
    """Render the fixed teams attendance page"""
    cities = City.objects.all()
    admin_password = settings.ADMIN_PASSWORD if hasattr(settings, 'ADMIN_PASSWORD') else "admin123"
    return render(request, 'staff_attendance/fixed_teams.html', {
        'cities': cities,
        'admin_password': admin_password,
    })


@csrf_exempt
@login_required
def get_fixed_teams_staff(request):
    """API endpoint to get the fixed teams staff with attendance status"""
    today = timezone.now().date()

    # Define fixed teams
    fixed_teams = {
        'contact': [
            'BENSAID HOUCINE',
            'AMRAOU MOHAMED',
            'AMRAOU ABDELLAH',
            'AMRAOU KHALED',
            'AMRAOU YOUSSEF',
            'BEN SAID LAHCEN',
            'BENDJEDDOU NADJIM',
            'BOUCHAIB TALSSI',
            'KHAMER TAWFIQ',
            'LOUNIS MAATI',
            'YEZLI AZIZ',
            'SALIHI AFRIM',
            'HAMAD MORAD',
            'SAKER FADI',
            'ZEROUALI MOHAMED',
            'QAZIMI ARBEN',
            'AIT EL GHACHI AHMED',
            'BOUYAHIAOUI ALI',
            'AMZIL KHALID',
            'DI CAMILLO FREDERIC',
            'MOUDEBER NADGEM',
            'DI SALVO LUCAS',
            'DI SALVO PATRICK',
            'MARONGIU THEO',
            'PARCOT MARC',
            'BOUADLA AMAR',
            'BULUT AHMET',
            'CADIR AYDIN',
            'OUAJOUB LAHCEN',
            'NDREJAJ RAMAZAN',
            'OUADAH ABDELREZAC',
            'BAGHA SAID',
            'NEAGU LAURENTIU',
            'KHAWJA HIJRATULLAH',
            'MAZOUJI RACHID',
            'RAI JAMAL',
            'RAI MOHAMED',
            'EL FIRARI KARIM',
            'YALCIN UMIT',
            'AMRAOU MOHAMED 2',
            'ZIMMERMANN LOIC',
        ],
        'moselle-est': [
            'BECKER SEBASTIEN',
            'FRANCOIS LAURENT',
            'KHOUYA MOHAMED',
            'REMIATTE LIONEL',
            'BOULIHSSAN YOUNES',
            'BENNACER ANISSA',
            'ANDRASCHKE DAVINA',
            'GUEZZI NAIMA',
            'BICER ROJDA',
            'KEHILI RANIA',
            'TAGHOULT-OUNMIR AMRAN',
        ],
        'maintenance': [
            'BOULAKDOUR YASSINE',
            'BEJENARU ANATOLIE',
            'AIT BAHA AHMED',
            'ADDI AHMED',
            'AIT ALLA SOFIENE',
            'OURAMI RACHID',
            'BRANGIER ROMUALD',
            'MAGAZ LAHCEN',
            'AGHOURI AREJDAL AISSA',
            'AMRI HATEM',
            'DARTE LOKMANE',
            'TAHARDJI HAMZA',
            'AIT BAHA YOUSSEF'

        ]
    }

    # Get today's attendance records for all staff WITH USER INFO
    attendances = Attendance.objects.filter(date=today).select_related('created_by',
                                                                       'updated_by')  # ← ADD select_related
    attendance_dict = {att.staff_member.name: att for att in attendances}

    # Check for active congé reservations
    active_conge = CongeReservation.objects.filter(
        start_date__lte=today,
        end_date__gte=today
    ).values_list('staff_member__name', flat=True)

    teams_data = {}

    for team_name, team_members in fixed_teams.items():
        team_data = []

        for member_name in team_members:
            # Get or create staff member (keep your existing logic)
            try:
                staff_member = StaffMember.objects.get(name=member_name)
            except StaffMember.DoesNotExist:
                # Create logic (keep your existing code)
                pass

            attendance = attendance_dict.get(staff_member.name)

            # Determine status (keep your existing logic)
            if staff_member.name in active_conge:
                status = 'conge'
            elif attendance:
                if attendance.absence_reason == 'CONGE_STATUS':
                    status = 'conge'
                elif attendance.present is True:
                    status = 'present'
                elif attendance.present is False:
                    status = 'absent'
                else:
                    status = 'undefined'
            else:
                status = 'undefined'

            # ADD THIS: Get user info for who made the last change
            last_modified_by = None
            if attendance:
                if attendance.updated_by:
                    last_modified_by = f"{attendance.updated_by.first_name or attendance.updated_by.username}"
                elif attendance.created_by:
                    last_modified_by = f"{attendance.created_by.first_name or attendance.created_by.username}"

            team_data.append({
                'id': staff_member.id,
                'name': staff_member.name,
                'status': status,
                'current_chantier': staff_member.city.name if staff_member.city.name != 'Équipes Fixes' else None,
                'absence_reason': attendance.absence_reason if attendance and attendance.present is False and attendance.absence_reason != 'CONGE_STATUS' else None,
                'timestamp': attendance.timestamp.isoformat() if attendance and attendance.timestamp else None,
                'hours_worked': float(attendance.hours_worked) if attendance and attendance.hours_worked else None,
                'grand_deplacement': attendance.grand_deplacement if attendance else False,
                'last_modified_by': last_modified_by  # ← ADD THIS LINE
            })

        teams_data[team_name] = team_data

    # Get all available chantiers for assignment (keep existing)
    chantiers = City.objects.exclude(name='Équipes Fixes').values('id', 'name')

    return JsonResponse({
        'teams': teams_data,
        'chantiers': list(chantiers)
    })

@csrf_exempt
@login_required  # Make sure this decorator is present!
def mark_fixed_team_present(request):
    """API endpoint to mark a fixed team member as present with chantier assignment and USER TRACKING"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')
        hours_worked = data.get('hours_worked', 8.0)
        grand_deplacement = data.get('grand_deplacement', False)
        chantier_id = data.get('chantier_id')

        try:
            staff = StaffMember.objects.get(id=staff_id)
            today = timezone.now().date()

            # Update staff member's chantier if provided
            if chantier_id:
                chantier = City.objects.get(id=chantier_id)
                staff.city = chantier
                staff.save()

            # Get or create attendance record
            attendance, created = Attendance.objects.get_or_create(
                staff_member=staff,
                date=today,
                defaults={
                    'present': True,
                    'timestamp': timezone.now(),
                    'hours_worked': hours_worked,
                    'grand_deplacement': grand_deplacement,
                    'created_by': request.user  # ADD: Track who created it
                }
            )

            # Update if already exists
            if not created:
                attendance.present = True
                attendance.absence_reason = None
                attendance.timestamp = timezone.now()
                attendance.hours_worked = hours_worked
                attendance.grand_deplacement = grand_deplacement
                attendance.updated_by = request.user  # ADD: Track who updated it
                attendance.save()

            chantier_text = f" sur {chantier.name}" if chantier_id else ""
            message = f"{staff.name} marqué présent ({hours_worked}h){chantier_text}"
            if grand_deplacement:
                message += " - Grand déplacement (nuit)"
            message += f" à {timezone.now().strftime('%H:%M:%S')}"

            return JsonResponse({
                'success': True,
                'message': message
            })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except City.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Chantier non trouvé'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required  # Make sure this decorator is present!
def mark_fixed_team_absent(request):
    """API endpoint to mark a fixed team member as absent with reason and USER TRACKING"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')
        reason = data.get('reason', '')

        try:
            staff = StaffMember.objects.get(id=staff_id)
            today = timezone.now().date()

            # Get or create attendance record
            attendance, created = Attendance.objects.get_or_create(
                staff_member=staff,
                date=today,
                defaults={
                    'present': False,
                    'timestamp': timezone.now(),
                    'absence_reason': reason,
                    'created_by': request.user  # ADD: Track who created it
                }
            )

            # Update if already exists
            if not created:
                attendance.present = False
                attendance.timestamp = timezone.now()
                attendance.absence_reason = reason
                attendance.updated_by = request.user  # ADD: Track who updated it
                attendance.save()

            return JsonResponse({
                'success': True,
                'message': f"{staff.name} marqué absent à {timezone.now().strftime('%H:%M:%S')}"
            })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)




@csrf_exempt
@login_required  # Make sure this decorator is present!
def mark_fixed_team_conge(request):
    """API endpoint to mark a fixed team member as on leave (congé) for today only with USER TRACKING"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')

        try:
            staff = StaffMember.objects.get(id=staff_id)
            today = timezone.now().date()

            # Get or create attendance record
            attendance, created = Attendance.objects.get_or_create(
                staff_member=staff,
                date=today,
                defaults={
                    'present': None,
                    'timestamp': timezone.now(),
                    'absence_reason': 'CONGE_STATUS',
                    'created_by': request.user  # ADD: Track who created it
                }
            )

            # Update if already exists
            if not created:
                attendance.present = None
                attendance.absence_reason = 'CONGE_STATUS'
                attendance.timestamp = timezone.now()
                attendance.updated_by = request.user  # ADD: Track who updated it
                attendance.save()

            return JsonResponse({
                'success': True,
                'message': f"{staff.name} marqué en congé à {timezone.now().strftime('%H:%M:%S')}"
            })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def get_department_stats(request):
    """API endpoint to get department statistics"""
    try:
        today = timezone.now().date()

        # Department mapping
        DEPARTMENT_MAPPING = {
            'contact': ['160 - DROUOT_Réhabilitation de 821 logts'],  # Remplacez par vos chantiers Contact
            'maintenance': ['001 - Bureau'],  # Remplacez par vos chantiers Maintenance
            'moselle-est': ['015 - Atelier']
        }

        stats = {}

        for dept, cities in DEPARTMENT_MAPPING.items():
            # Get staff in this department
            staff_in_dept = StaffMember.objects.filter(city__name__in=cities)
            total_staff = staff_in_dept.count()

            # Get attendance for today
            attendances = Attendance.objects.filter(
                staff_member__in=staff_in_dept,
                date=today
            )

            present_count = attendances.filter(present=True).count()
            absent_count = attendances.filter(present=False).count()
            conge_count = attendances.filter(absence_reason='CONGE_STATUS').count()
            undefined_count = total_staff - present_count - absent_count - conge_count

            stats[dept] = {
                'total': total_staff,
                'present': present_count,
                'absent': absent_count,
                'conge': conge_count,
                'undefined': undefined_count,
                'cities': cities
            }

        return JsonResponse({
            'success': True,
            'stats': stats,
            'date': today.isoformat()
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def update_city_name(request):
    """API endpoint to update a city's name"""
    if request.method == 'POST':
        data = json.loads(request.body)
        city_id = data.get('city_id')
        new_name = data.get('new_name', '').strip()

        if not new_name:
            return JsonResponse({'success': False, 'message': 'Le nouveau nom est requis'}, status=400)

        try:
            city = City.objects.get(id=city_id)
            old_name = city.name
            city.name = new_name
            city.save()

            return JsonResponse({
                'success': True,
                'message': f'Ville renommée de "{old_name}" à "{new_name}"',
                'city': {
                    'id': city.id,
                    'name': city.name
                }
            })
        except City.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Ville non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def delete_city(request):
    """API endpoint to delete a city"""
    if request.method == 'POST':
        data = json.loads(request.body)
        city_id = data.get('city_id')

        try:
            city = City.objects.get(id=city_id)

            # Check if city has staff members
            staff_count = StaffMember.objects.filter(city=city).count()
            if staff_count > 0:
                return JsonResponse({
                    'success': False,
                    'message': f'Impossible de supprimer "{city.name}". Elle contient {staff_count} membre(s) du personnel.'
                }, status=400)

            city_name = city.name
            city.delete()

            return JsonResponse({
                'success': True,
                'message': f'Ville "{city_name}" supprimée avec succès'
            })
        except City.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Ville non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def mark_conge_with_period(request):
    """API endpoint to mark a staff member as on leave with user tracking"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        reason = data.get('reason', '')

        try:
            staff = StaffMember.objects.get(id=staff_id)

            # Parse dates
            start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()

            # Validate dates
            if start_date_obj > end_date_obj:
                return JsonResponse(
                    {'success': False, 'message': 'La date de début doit être antérieure à la date de fin'},
                    status=400
                )

            # Check for overlapping congé periods
            overlapping = CongeReservation.objects.filter(
                staff_member=staff,
                start_date__lte=end_date_obj,
                end_date__gte=start_date_obj
            ).exists()

            if overlapping:
                return JsonResponse(
                    {'success': False, 'message': 'Cette période de congé chevauche avec une période existante'},
                    status=400
                )

            # Create congé reservation with user tracking
            conge = CongeReservation.objects.create(
                staff_member=staff,
                start_date=start_date_obj,
                end_date=end_date_obj,
                reason=reason,
                created_by=request.user  # Track who created it
            )

            message = f"{staff.name} en congé du {start_date_obj.strftime('%d/%m/%Y')} au {end_date_obj.strftime('%d/%m/%Y')} (créé par {request.user.first_name or request.user.username})"

            return JsonResponse({
                'success': True,
                'message': message
            })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Format de date invalide'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def create_user(request):
    """API endpoint to create a new user (admin only)"""
    profile = UserProfile.objects.get_or_create(user=request.user)[0]
    if not profile.is_admin:
        return JsonResponse({'success': False, 'message': 'Accès refusé'}, status=403)

    if request.method == 'POST':
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        first_name = data.get('first_name', '')
        last_name = data.get('last_name', '')
        role = data.get('role', 'user')
        can_modify_all = data.get('can_modify_all', False)
        assigned_zones = data.get('assigned_zones', [])

        if not username or not password:
            return JsonResponse({'success': False, 'message': 'Nom d\'utilisateur et mot de passe requis'}, status=400)

        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'message': 'Ce nom d\'utilisateur existe déjà'}, status=400)

        try:
            # Create user
            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=first_name,
                last_name=last_name
            )

            # Create profile
            user_profile = UserProfile.objects.create(
                user=user,
                role=role,
                can_modify_all=can_modify_all
            )

            # Assign zones
            if assigned_zones:
                zones = Zone.objects.filter(id__in=assigned_zones)
                user_profile.assigned_zones.set(zones)

            return JsonResponse({
                'success': True,
                'message': f'Utilisateur {username} créé avec succès',
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'role': user_profile.role,
                    'can_modify_all': user_profile.can_modify_all,
                    'assigned_zones': list(user_profile.assigned_zones.values_list('name', flat=True))
                }
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def get_current_user(request):
    """API endpoint to get current user info"""
    profile = UserProfile.objects.get_or_create(user=request.user)[0]

    return JsonResponse({
        'success': True,
        'user': {
            'id': request.user.id,
            'username': request.user.username,
            'first_name': request.user.first_name,
            'last_name': request.last_name,
            'role': profile.role,
            'is_admin': profile.is_admin,
            'can_modify_all': profile.can_modify_all,
            'assigned_zones': list(profile.assigned_zones.values_list('name', flat=True))
        }
    })


@csrf_exempt
def get_conge_reservations(request):
    """API endpoint to get current and future congé reservations + today's congés"""
    try:
        today = timezone.now().date()

        # Get congé reservations (périodes)
        reservations = CongeReservation.objects.filter(
            end_date__gte=today  # Only show if end date is today or in the future
        ).select_related('staff_member__city').order_by('start_date')

        # Get today's single-day congés (from drag & drop)
        todays_conges = Attendance.objects.filter(
            date=today,
            absence_reason='CONGE_STATUS'
        ).exclude(
            # Exclude those who already have a reservation for today
            staff_member__conge_reservations__start_date__lte=today,
            staff_member__conge_reservations__end_date__gte=today
        ).select_related('staff_member__city')

        conge_data = []

        # Add reservation-based congés
        for reservation in reservations:
            conge_data.append({
                'id': f"reservation_{reservation.id}",
                'type': 'reservation',
                'reservation_id': reservation.id,
                'staff_id': reservation.staff_member.id,
                'staff_name': reservation.staff_member.name,
                'city': reservation.staff_member.city.name,
                'start_date': reservation.start_date.isoformat(),
                'end_date': reservation.end_date.isoformat(),
                'start_date_display': reservation.start_date.strftime('%d/%m/%Y'),
                'end_date_display': reservation.end_date.strftime('%d/%m/%Y'),
                'reason': reservation.reason,
                'is_current': reservation.start_date <= today <= reservation.end_date,
                'is_single_day': False
            })

        # Add single-day congés
        for attendance in todays_conges:
            conge_data.append({
                'id': f"daily_{attendance.staff_member.id}",
                'type': 'daily',
                'staff_id': attendance.staff_member.id,
                'staff_name': attendance.staff_member.name,
                'city': attendance.staff_member.city.name,
                'start_date': today.isoformat(),
                'end_date': today.isoformat(),
                'start_date_display': today.strftime('%d/%m/%Y'),
                'end_date_display': today.strftime('%d/%m/%Y'),
                'reason': None,
                'is_current': True,
                'is_single_day': True
            })

        # Sort by current status first, then by date
        conge_data.sort(key=lambda x: (not x['is_current'], x['start_date']))

        return JsonResponse({'success': True, 'reservations': conge_data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def remove_conge_reservation(request):
    """API endpoint to remove a congé reservation"""
    if request.method == 'POST':
        data = json.loads(request.body)
        reservation_id = data.get('reservation_id')

        try:
            reservation = CongeReservation.objects.get(id=reservation_id)
            staff_name = reservation.staff_member.name
            reservation.delete()  # This will also remove related attendance records

            return JsonResponse({
                'success': True,
                'message': f'Congé de {staff_name} supprimé avec succès'
            })
        except CongeReservation.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Réservation de congé non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def remove_daily_conge(request):
    """API endpoint to remove a single-day congé"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')

        try:
            today = timezone.now().date()
            attendance = Attendance.objects.get(
                staff_member_id=staff_id,
                date=today,
                absence_reason='CONGE_STATUS'
            )

            staff_name = attendance.staff_member.name
            attendance.delete()

            return JsonResponse({
                'success': True,
                'message': f'Congé du jour de {staff_name} supprimé avec succès'
            })
        except Attendance.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Congé du jour non trouvé'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def get_staff_list(request):
    """API endpoint to get the list of staff members with attendance status and user info"""
    today = timezone.now().date()

    # Get all staff members with their cities
    staff_members = StaffMember.objects.select_related('city').all()

    # Get today's attendance records with user info
    attendances = Attendance.objects.filter(date=today).select_related('created_by', 'updated_by')

    # Create a dictionary for quick lookup
    attendance_dict = {att.staff_member_id: att for att in attendances}

    # Check for active congé reservations
    active_conge = CongeReservation.objects.filter(
        start_date__lte=today,
        end_date__gte=today
    ).values_list('staff_member_id', flat=True)

    # Prepare response data
    staff_data = []
    for staff in staff_members:
        attendance = attendance_dict.get(staff.id)

        # Check if staff is in congé TODAY
        if staff.id in active_conge:
            status = 'conge'
        elif attendance:
            if attendance.absence_reason == 'CONGE_STATUS':
                status = 'conge'
            elif attendance.present is True:
                status = 'present'
            elif attendance.present is False:
                status = 'absent'
            else:
                status = 'undefined'
        else:
            status = 'undefined'

        # Get user info for who made the last change
        last_modified_by = None
        if attendance:
            if attendance.updated_by:
                last_modified_by = f"{attendance.updated_by.first_name or attendance.updated_by.username}"
            elif attendance.created_by:
                last_modified_by = f"{attendance.created_by.first_name or attendance.created_by.username}"

        staff_data.append({
            'id': staff.id,
            'name': staff.name,
            'city': staff.city.name,
            'status': status,
            'absence_reason': attendance.absence_reason if attendance and attendance.present is False and attendance.absence_reason != 'CONGE_STATUS' else None,
            'timestamp': attendance.timestamp.isoformat() if attendance and attendance.timestamp else None,
            'hours_worked': float(attendance.hours_worked) if attendance and attendance.hours_worked else None,
            'grand_deplacement': attendance.grand_deplacement if attendance else False,
            'last_modified_by': last_modified_by  # NEW: Who made the last change
        })

    return JsonResponse({'staffMembers': staff_data})


# Additional views to add to staff_attendance/views.py

@csrf_exempt
@login_required
def get_users_list(request):
    """API endpoint to get list of all users (admin only)"""
    profile = UserProfile.objects.get_or_create(user=request.user)[0]
    if not profile.is_admin:
        return JsonResponse({'success': False, 'message': 'Accès refusé'}, status=403)

    try:
        users = User.objects.all().select_related('profile').prefetch_related('profile__assigned_zones')
        users_data = []

        for user in users:
            user_profile = UserProfile.objects.get_or_create(user=user)[0]
            assigned_zones = list(user_profile.assigned_zones.values_list('name', flat=True))

            users_data.append({
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'role': user_profile.role,
                'can_modify_all': user_profile.can_modify_all,
                'assigned_zones': assigned_zones,
                'created_at': user_profile.created_at.isoformat(),
                'last_login': user.last_login.isoformat() if user.last_login else None
            })

        return JsonResponse({
            'success': True,
            'users': users_data
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required
def delete_user(request):
    """API endpoint to delete a user (admin only)"""
    profile = UserProfile.objects.get_or_create(user=request.user)[0]
    if not profile.is_admin:
        return JsonResponse({'success': False, 'message': 'Accès refusé'}, status=403)

    if request.method == 'POST':
        data = json.loads(request.body)
        user_id = data.get('user_id')

        if not user_id:
            return JsonResponse({'success': False, 'message': 'ID utilisateur requis'}, status=400)

        # Prevent self-deletion
        if user_id == request.user.id:
            return JsonResponse({'success': False, 'message': 'Vous ne pouvez pas supprimer votre propre compte'},
                                status=400)

        try:
            user = User.objects.get(id=user_id)
            username = user.username
            user.delete()

            return JsonResponse({
                'success': True,
                'message': f'Utilisateur "{username}" supprimé avec succès'
            })

        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Utilisateur non trouvé'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def get_attendance_with_users(request):
    """API endpoint to get attendance history with user information"""
    try:
        # Get month and year from request parameters
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not month or not year:
            today = timezone.now().date()
            month = today.month
            year = today.year
        else:
            month = int(month)
            year = int(year)

        # Calculate date range
        import calendar
        start_date = datetime.date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime.date(year, month, last_day)

        # Get attendance records with user information
        attendances = Attendance.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).select_related('staff_member__city', 'created_by', 'updated_by').order_by('-date')

        attendance_data = []
        for attendance in attendances:
            # Get the user who made the last change
            last_modified_by = None
            if attendance.updated_by:
                last_modified_by = f"{attendance.updated_by.first_name} {attendance.updated_by.last_name}".strip()
                if not last_modified_by:
                    last_modified_by = attendance.updated_by.username
            elif attendance.created_by:
                last_modified_by = f"{attendance.created_by.first_name} {attendance.created_by.last_name}".strip()
                if not last_modified_by:
                    last_modified_by = attendance.created_by.username

            attendance_data.append({
                'staff_name': attendance.staff_member.name,
                'city': attendance.staff_member.city.name,
                'date': attendance.date.isoformat(),
                'status': attendance.status,
                'hours_worked': float(attendance.hours_worked) if attendance.hours_worked else None,
                'grand_deplacement': attendance.grand_deplacement,
                'absence_reason': attendance.absence_reason if attendance.present is False and attendance.absence_reason != 'CONGE_STATUS' else None,
                'timestamp': attendance.timestamp.isoformat() if attendance.timestamp else None,
                'last_modified_by': last_modified_by
            })

        return JsonResponse({
            'success': True,
            'attendance': attendance_data,
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month]
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
# Add this to your views.py if you want an API endpoint to clean old congés
@csrf_exempt
@login_required
def get_departments_data(request):
    """API endpoint to get staff organized by departments with user tracking - FINAL VERSION"""
    try:
        today = timezone.now().date()

        # Fixed teams mapping
        DEPARTMENT_MAPPING = {
            'contact': [
                'BENSAID HOUCINE', 'AMRAOU MOHAMED', 'AMRAOU ABDELLAH', 'AMRAOU KHALED',
                'AMRAOU YOUSSEF', 'BEN SAID LAHCEN', 'BENDJEDDOU NADJIM', 'BOUCHAIB TALSSI',
                'KHAMER TAWFIQ', 'LOUNIS MAATI', 'YEZLI AZIZ', 'SALIHI AFRIM',
                'HAMAD MORAD', 'SAKER FADI', 'ZEROUALI MOHAMED', 'QAZIMI ARBEN',
                'AIT EL GHACHI AHMED', 'BOUYAHIAOUI ALI', 'AMZIL KHALID', 'DI CAMILLO FREDERIC',
                'MOUDEBER NADGEM', 'DI SALVO LUCAS', 'DI SALVO PATRICK', 'MARONGIU THEO',
                'PARCOT MARC', 'BOUADLA AMAR', 'BULUT AHMET', 'CADIR AYDIN',
                'OUAJOUB LAHCEN', 'NDREJAJ RAMAZAN', 'OUADAH ABDELREZAC', 'BAGHA SAID',
                'NEAGU LAURENTIU', 'KHAWJA HIJRATULLAH', 'MAZOUJI RACHID', 'RAI JAMAL',
                'RAI MOHAMED', 'EL FIRARI KARIM', 'YALCIN UMIT', 'AMRAOU MOHAMED 2',
                'ZIMMERMANN LOIC'
            ],
            'moselle-est': [
                'BECKER SEBASTIEN', 'FRANCOIS LAURENT', 'KHOUYA MOHAMED', 'REMIATTE LIONEL',
                'BOULIHSSAN YOUNES', 'BENNACER ANISSA', 'ANDRASCHKE DAVINA', 'GUEZZI NAIMA',
                'BICER ROJDA', 'KEHILI RANIA', 'TAGHOULT-OUNMIR AMRAN'
            ],
            'maintenance': [
                'BOULAKDOUR YASSINE', 'BEJENARU ANATOLIE', 'AIT BAHA AHMED', 'ADDI AHMED',
                'AIT ALLA SOFIENE', 'OURAMI RACHID', 'BRANGIER ROMUALD', 'MAGAZ LAHCEN',
                'AGHOURI AREJDAL AISSA', 'AMRI HATEM', 'DARTE LOKMANE', 'TAHARDJI HAMZA', 'AIT BAHA YOUSSEF'
            ]
        }

        # Get today's attendance records WITH user info
        attendances = Attendance.objects.filter(date=today).select_related('staff_member', 'created_by', 'updated_by')
        attendance_dict = {att.staff_member_id: att for att in attendances}

        # Check for active congé reservations
        active_conge = CongeReservation.objects.filter(
            start_date__lte=today,
            end_date__gte=today
        ).values_list('staff_member_id', flat=True)

        # Organize staff by departments
        departments = {
            'contact': [],
            'moselle-est': [],
            'maintenance': []
        }

        # Process each department
        for dept_name, member_names in DEPARTMENT_MAPPING.items():
            for member_name in member_names:
                try:
                    staff_member = StaffMember.objects.get(name=member_name)
                    # Skip if contract has expired
                    if staff_member.contract_status == 'expire':
                        continue  # Ne pas inclure ce membre dans la liste
                except StaffMember.DoesNotExist:
                    default_city, created = City.objects.get_or_create(
                        name='Équipes Fixes',
                        defaults={'zone': None}
                    )
                    staff_member = StaffMember.objects.create(
                        name=member_name,
                        city=default_city
                    )

                attendance = attendance_dict.get(staff_member.id)

                # Determine status
                if staff_member.id in active_conge:
                    status = 'conge'
                elif attendance:
                    if attendance.absence_reason == 'CONGE_STATUS':
                        status = 'conge'
                    elif attendance.present is True:
                        status = 'present'
                    elif attendance.present is False:
                        status = 'absent'
                    else:
                        status = 'undefined'
                else:
                    status = 'undefined'

                # Get user info for who made the last change
                last_modified_by = None
                if attendance:
                    if attendance.updated_by:
                        last_modified_by = f"{attendance.updated_by.first_name or attendance.updated_by.username}"
                    elif attendance.created_by:
                        last_modified_by = f"{attendance.created_by.first_name or attendance.created_by.username}"

                # Create staff data
                staff_data = {
                    'id': staff_member.id,
                    'name': staff_member.name,
                    'status': status,
                    'current_chantier': staff_member.city.name if staff_member.city.name != 'Équipes Fixes' else None,
                    'absence_reason': attendance.absence_reason if attendance and attendance.present is False and attendance.absence_reason != 'CONGE_STATUS' else None,
                    'timestamp': attendance.timestamp.isoformat() if attendance and attendance.timestamp else None,
                    'hours_worked': float(attendance.hours_worked) if attendance and attendance.hours_worked else None,
                    'grand_deplacement': attendance.grand_deplacement if attendance else False,
                    'last_modified_by': last_modified_by
                }

                departments[dept_name].append(staff_data)

        # Get all available chantiers for assignment
        chantiers = City.objects.exclude(name='Équipes Fixes').values('id', 'name')

        return JsonResponse({
            'success': True,
            'departments': departments,
            'chantiers': list(chantiers),
            'date': today.isoformat()
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required
def get_contract_alerts(request):
    """API endpoint to get contract expiration alerts"""
    try:
        today = timezone.now().date()

        # Staff avec contrats qui expirent dans 7 jours ou moins
        warning_contracts = StaffMember.objects.filter(
            date_fin_contrat__isnull=False,
            date_fin_contrat__gte=today,  # Pas encore expiré
            date_fin_contrat__lte=today + timezone.timedelta(days=7)  # Expire dans 7 jours max
        ).select_related('city')

        # Staff avec contrats expirés (pour nettoyage éventuel)
        expired_contracts = StaffMember.objects.filter(
            date_fin_contrat__isnull=False,
            date_fin_contrat__lt=today
        ).select_related('city')

        alerts = []

        for staff in warning_contracts:
            days_remaining = staff.days_until_contract_end
            alerts.append({
                'id': staff.id,
                'name': staff.name,
                'city': staff.city.name,
                'date_fin_contrat': staff.date_fin_contrat.isoformat(),
                'days_remaining': days_remaining,
                'status': 'warning',
                'message': f"Le contrat de {staff.name} expire dans {days_remaining} jour{'s' if days_remaining > 1 else ''} ({staff.date_fin_contrat.strftime('%d/%m/%Y')})"
            })

        expired_list = []
        for staff in expired_contracts:
            expired_list.append({
                'id': staff.id,
                'name': staff.name,
                'city': staff.city.name,
                'date_fin_contrat': staff.date_fin_contrat.isoformat(),
                'days_expired': abs(staff.days_until_contract_end),
                'status': 'expired'
            })

        return JsonResponse({
            'success': True,
            'alerts': alerts,
            'expired': expired_list,
            'warning_count': len(alerts),
            'expired_count': len(expired_list)
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required
def cleanup_expired_contracts(request):
    """API endpoint to remove staff with expired contracts"""
    if request.method == 'POST':
        try:
            today = timezone.now().date()

            expired_staff = StaffMember.objects.filter(
                date_fin_contrat__isnull=False,
                date_fin_contrat__lt=today
            )

            expired_names = [staff.name for staff in expired_staff]
            count = expired_staff.count()

            if count > 0:
                # Supprimer les enregistrements de présence liés (optionnel)
                # Attendance.objects.filter(staff_member__in=expired_staff).delete()

                # Supprimer le personnel
                expired_staff.delete()

                return JsonResponse({
                    'success': True,
                    'message': f'{count} membre(s) du personnel avec contrats expirés supprimé(s)',
                    'removed_staff': expired_names,
                    'count': count
                })
            else:
                return JsonResponse({
                    'success': True,
                    'message': 'Aucun contrat expiré à nettoyer',
                    'count': 0
                })

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)

def home_redirect(request):
    """Handle root URL redirect based on authentication status"""
    if request.user.is_authenticated:
        return redirect('staff_attendance:attendance_page')
    else:
        return redirect('staff_attendance:login')

@csrf_exempt
def cleanup_past_conges(request):
    """API endpoint to clean up past congé reservations"""
    if request.method == 'POST':
        try:
            days = request.POST.get('days', 30)  # Default 30 days
            days = int(days)

            cutoff_date = timezone.now().date() - timezone.timedelta(days=days)

            # Find and delete past congés
            past_conges = CongeReservation.objects.filter(end_date__lt=cutoff_date)
            count = past_conges.count()

            if count > 0:
                past_conges.delete()

            return JsonResponse({
                'success': True,
                'message': f'{count} anciens congés supprimés',
                'cleaned_count': count
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def get_attendance_history(request):
    """API endpoint to get attendance history for a specific month/year"""
    try:
        # Get month and year from request parameters
        month = request.GET.get('month')
        year = request.GET.get('year')

        # If no month/year provided, use current month
        if not month or not year:
            today = timezone.now().date()
            month = today.month
            year = today.year
        else:
            month = int(month)
            year = int(year)

        # Calculate the first and last day of the requested month
        import calendar
        start_date = datetime.date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime.date(year, month, last_day)

        # Get all staff members with their cities
        staff_members = StaffMember.objects.select_related('city').all()

        # Get attendance records for the requested month
        attendances = Attendance.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).order_by('-date')

        # Group attendances by date
        attendance_by_date = {}

        # Initialize all dates in the month with empty dictionaries
        current_date = start_date
        while current_date <= end_date:
            attendance_by_date[current_date.isoformat()] = {
                'date': current_date.isoformat(),
                'display_date': current_date.strftime('%d/%m/%Y'),
                'weekday': current_date.strftime('%A'),
                'staff': {}
            }
            current_date += datetime.timedelta(days=1)

        # Fill in attendance data
        for attendance in attendances:
            date_key = attendance.date.isoformat()
            if date_key in attendance_by_date:
                if attendance.absence_reason == 'CONGE_STATUS':
                    status = 'conge'
                elif attendance.present is True:
                    status = 'present'
                elif attendance.present is False:
                    status = 'absent'
                else:
                    status = 'undefined'

                attendance_by_date[date_key]['staff'][attendance.staff_member_id] = {
                    'status': status,
                    'absence_reason': attendance.absence_reason if attendance.present is False and attendance.absence_reason != 'CONGE_STATUS' else None,
                    'timestamp': attendance.timestamp.isoformat() if attendance.timestamp else None,
                    'hours_worked': float(attendance.hours_worked) if attendance.hours_worked else None,
                    'grand_deplacement': attendance.grand_deplacement if attendance.grand_deplacement else False
                }

        # Convert to a list and sort by date (newest first)
        history_data = list(attendance_by_date.values())
        history_data.sort(key=lambda x: x['date'])

        # Prepare staff info
        staff_info = []
        for staff in staff_members:
            staff_info.append({
                'id': staff.id,
                'name': staff.name,
                'city': staff.city.name
            })

        # Calculate some statistics for the month
        total_days = len(history_data)
        working_days = len([day for day in history_data if day['weekday'] not in ['Saturday', 'Sunday']])

        return JsonResponse({
            'history': history_data,
            'staff': staff_info,
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month],
            'total_days': total_days,
            'working_days': working_days,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        })

    except ValueError as e:
        return JsonResponse({'success': False, 'message': 'Mois ou année invalide'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def get_staff_by_city(request):
    """API endpoint to get staff members by city"""
    if request.method == 'GET':
        city_id = request.GET.get('city_id')

        if city_id:
            try:
                city = City.objects.get(id=city_id)
                staff_members = StaffMember.objects.filter(city=city)

                staff_data = []
                for staff in staff_members:
                    staff_data.append({
                        'id': staff.id,
                        'name': staff.name,
                        'city_id': staff.city.id,
                        'city_name': staff.city.name
                    })

                return JsonResponse({
                    'success': True,
                    'staff': staff_data,
                    'city_name': city.name
                })
            except City.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Ville non trouvée'}, status=404)
        else:
            # Return all staff if no city specified
            staff_members = StaffMember.objects.select_related('city').all()
            staff_data = []
            for staff in staff_members:
                staff_data.append({
                    'id': staff.id,
                    'name': staff.name,
                    'city_id': staff.city.id,
                    'city_name': staff.city.name
                })

            return JsonResponse({
                'success': True,
                'staff': staff_data
            })

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def mark_present(request):
    """Make sure this sets the user tracking fields"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')
        hours_worked = data.get('hours_worked', 8.0)
        grand_deplacement = data.get('grand_deplacement', False)

        try:
            staff = StaffMember.objects.get(id=staff_id)
            today = timezone.now().date()

            print(f"DEBUG: Marking {staff.name} present by user {request.user}")

            # Get or create attendance record
            attendance, created = Attendance.objects.get_or_create(
                staff_member=staff,
                date=today,
                defaults={
                    'present': True,
                    'timestamp': timezone.now(),
                    'hours_worked': hours_worked,
                    'grand_deplacement': grand_deplacement,
                    'created_by': request.user  # Make sure this is set!
                }
            )

            # Update if already exists
            if not created:
                print(f"DEBUG: Updating existing attendance, setting updated_by to {request.user}")
                attendance.present = True
                attendance.absence_reason = None
                attendance.timestamp = timezone.now()
                attendance.hours_worked = hours_worked
                attendance.grand_deplacement = grand_deplacement
                attendance.updated_by = request.user  # Make sure this is set!
                attendance.save()
            else:
                print(f"DEBUG: Created new attendance with created_by = {request.user}")

            return JsonResponse({
                'success': True,
                'message': f"{staff.name} marqué présent"
            })

        except Exception as e:
            print(f"DEBUG: Error in mark_present: {e}")
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)



@csrf_exempt
@login_required
def mark_absent(request):
    """API endpoint to mark a staff member as absent with user tracking"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')
        reason = data.get('reason', '')

        try:
            staff = StaffMember.objects.get(id=staff_id)
            today = timezone.now().date()

            # Get or create attendance record
            attendance, created = Attendance.objects.get_or_create(
                staff_member=staff,
                date=today,
                defaults={
                    'present': False,
                    'timestamp': timezone.now(),
                    'absence_reason': reason,
                    'created_by': request.user
                }
            )

            # Update if already exists
            if not created:
                attendance.present = False
                attendance.timestamp = timezone.now()
                attendance.absence_reason = reason
                attendance.updated_by = request.user
                attendance.save()

            message = f"{staff.name} marqué absent à {timezone.now().strftime('%H:%M:%S')} par {request.user.first_name or request.user.username}"

            return JsonResponse({
                'success': True,
                'message': message
            })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)





@csrf_exempt
def mark_conge(request):
    """API endpoint to mark a staff member as on leave (congé) for today only"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')

        try:
            staff = StaffMember.objects.get(id=staff_id)
            today = timezone.now().date()

            # Get or create attendance record
            attendance, created = Attendance.objects.get_or_create(
                staff_member=staff,
                date=today,
                defaults={'present': None, 'timestamp': timezone.now(), 'absence_reason': 'CONGE_STATUS'}
            )

            # Update if already exists
            if not created:
                attendance.present = None  # NULL for congé status
                attendance.absence_reason = 'CONGE_STATUS'  # Special identifier for congé
                attendance.timestamp = timezone.now()
                attendance.save()

            return JsonResponse({
                'success': True,
                'message': f"{staff.name} marqué en congé à {timezone.now().strftime('%H:%M:%S')}"
            })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def mark_city_present(request):
    """API endpoint to mark all staff members of a city as present with default values"""
    if request.method == 'POST':
        data = json.loads(request.body)
        city_name = data.get('city')
        default_hours = data.get('default_hours', 8.0)  # Default 8 hours
        default_grand_deplacement = data.get('default_grand_deplacement', False)  # Default no

        try:
            city = City.objects.get(name=city_name)
            staff_members = StaffMember.objects.filter(city=city)
            today = timezone.now().date()
            current_time = timezone.now()

            # Mark all staff in the city as present with default values
            for staff in staff_members:
                attendance, created = Attendance.objects.get_or_create(
                    staff_member=staff,
                    date=today,
                    defaults={
                        'present': True,
                        'timestamp': current_time,
                        'hours_worked': default_hours,
                        'grand_deplacement': default_grand_deplacement
                    }
                )

                if not created:
                    attendance.present = True
                    attendance.absence_reason = None  # Clear any absence reason
                    attendance.timestamp = current_time
                    attendance.hours_worked = default_hours
                    attendance.grand_deplacement = default_grand_deplacement
                    attendance.save()

            grand_deplacement_text = " avec grand déplacement" if default_grand_deplacement else ""
            return JsonResponse({
                'success': True,
                'message': f"Tout le personnel de {city_name} marqué présent ({default_hours}h{grand_deplacement_text}) !"
            })

        except City.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Ville non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def mark_all_present(request):
    """API endpoint to mark all staff members as present"""
    if request.method == 'POST':
        try:
            staff_members = StaffMember.objects.all()
            today = timezone.now().date()
            current_time = timezone.now()

            # Mark all staff as present
            for staff in staff_members:
                attendance, created = Attendance.objects.get_or_create(
                    staff_member=staff,
                    date=today,
                    defaults={'present': True, 'timestamp': current_time}
                )

                if not created:
                    attendance.present = True
                    attendance.absence_reason = None  # Clear any absence reason
                    attendance.timestamp = current_time
                    attendance.save()

            return JsonResponse({
                'success': True,
                'message': "Tout le personnel marqué présent !"
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def update_staff_city(request):
    """API endpoint to update staff member's city"""
    if request.method == 'POST':
        data = json.loads(request.body)
        staff_id = data.get('staff_id')
        new_city_id = data.get('city_id')
        new_city_name = data.get('city_name')

        try:
            staff = StaffMember.objects.get(id=staff_id)

            # Handle both city_id and city_name parameters
            if new_city_id:
                new_city = City.objects.get(id=new_city_id)
            elif new_city_name:
                new_city = City.objects.get(name=new_city_name)
            else:
                return JsonResponse({'success': False, 'message': 'City ID ou nom de ville requis'}, status=400)

            old_city_name = staff.city.name
            staff.city = new_city
            staff.save()

            return JsonResponse({
                'success': True,
                'message': f"{staff.name} déplacé de {old_city_name} vers {new_city.name}"
            })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except City.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Ville non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def verify_admin(request):
    """API endpoint to verify admin password"""
    if request.method == 'POST':
        data = json.loads(request.body)
        password = data.get('password')

        # In a real application, use a more secure authentication method
        admin_password = settings.ADMIN_PASSWORD if hasattr(settings, 'ADMIN_PASSWORD') else "admin123"

        if password == admin_password:
            return JsonResponse({'success': True, 'message': 'Mode administrateur activé'})
        else:
            return JsonResponse({'success': False, 'message': 'Mot de passe incorrect'}, status=401)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


# Fonctions PDF améliorées et simplifiées
# À remplacer dans votre views.py

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from io import BytesIO
import calendar
import locale

# Dictionnaire des mois en français
MOIS_FRANCAIS = {
    1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
    5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
    9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
}

# Dictionnaire des jours en français
JOURS_FRANCAIS = {
    'Monday': 'Lundi', 'Tuesday': 'Mardi', 'Wednesday': 'Mercredi',
    'Thursday': 'Jeudi', 'Friday': 'Vendredi', 'Saturday': 'Samedi', 'Sunday': 'Dimanche'
}


@csrf_exempt
def generate_pdf(request):
    """API endpoint to generate a daily PDF report of attendance in landscape format"""
    if request.method == 'POST':
        try:
            # Get today's date
            today = timezone.now().date()

            # Get all staff with their attendance
            staff_members = StaffMember.objects.select_related('city').all()

            # Get attendance records for today
            attendances = Attendance.objects.filter(date=today)
            attendance_dict = {att.staff_member_id: att for att in attendances}

            # Create a BytesIO buffer for the PDF
            buffer = BytesIO()

            # Create the PDF object in LANDSCAPE mode
            from reportlab.lib.pagesizes import A4, landscape
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),  # Landscape for better table visibility
                topMargin=30,
                bottomMargin=30,
                leftMargin=30,
                rightMargin=30
            )
            elements = []

            # Set up styles
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']

            # Add title
            elements.append(Paragraph("Rapport de Présence du Personnel", title_style))
            elements.append(Spacer(1, 12))

            # Add date
            date_str = today.strftime("%A %d %B %Y")
            date_str = date_str.capitalize()  # Capitalize first letter
            elements.append(Paragraph(date_str, subtitle_style))
            elements.append(Spacer(1, 12))

            # Create table data with expanded columns for landscape
            data = [
                ["Nom du Personnel", "Chantier", "Statut", "Heures Travaillées", "Grand Déplacement",
                 "Motif d'Absence"]]


            # Add staff data to table
            for staff in staff_members:
                attendance = attendance_dict.get(staff.id)

                # Determine status
                if attendance:
                    if attendance.absence_reason == 'CONGE_STATUS':
                        status = "En congé"
                    elif attendance.present is True:
                        status = "Présent"
                    elif attendance.present is False:
                        status = "Absent"
                    else:
                        status = "Non défini"
                else:
                    status = "Non défini"



                # Get absence reason
                absence_reason = ""
                if attendance and attendance.present is False and attendance.absence_reason:
                    absence_reason = attendance.absence_reason

                # NOUVELLE LOGIQUE : Chantier vide si absent ou en congé
                chantier_name = ""
                if status == "Présent":
                    chantier_name = staff.city.name
                # Si absent ou en congé, chantier_name reste vide
                hours_worked = ""
                if attendance and attendance.hours_worked and attendance.present is True:
                    # Convert Decimal to float first
                    hours_float = float(attendance.hours_worked)
                    # Check if it's a whole number
                    if hours_float.is_integer():
                        hours = int(hours_float)
                    else:
                        hours = hours_float
                    hours_worked = f"{hours}h"
                grand_deplacement = ""
                if attendance and attendance.grand_deplacement and attendance.present is True:
                    grand_deplacement = "Oui"
                data.append([
                    staff.name,
                    chantier_name,  # ← MODIFICATION ICI
                    status,
                    hours_worked,
                    grand_deplacement,

                    absence_reason
                ])

            # Create table with optimized column widths for landscape
            # Total width available in landscape A4 minus margins ≈ 760 points
            table = Table(data, colWidths=[140, 100, 80, 90, 90, 80, 180])

            # Style the table
            table_style = TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

                # Body styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

                # Left align text columns
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Name
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # City
                ('ALIGN', (6, 1), (6, -1), 'LEFT'),  # Absence reason

                # Alternate row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ])

            # Add conditional formatting for status column
            for i in range(1, len(data)):
                status = data[i][2]
                if status == "Présent":
                    table_style.add('BACKGROUND', (2, i), (2, i), colors.lightgreen)
                    table_style.add('TEXTCOLOR', (2, i), (2, i), colors.darkgreen)
                    table_style.add('FONTNAME', (2, i), (2, i), 'Helvetica-Bold')
                elif status == "Absent":
                    table_style.add('BACKGROUND', (2, i), (2, i), colors.lightcoral)
                    table_style.add('TEXTCOLOR', (2, i), (2, i), colors.darkred)
                    table_style.add('FONTNAME', (2, i), (2, i), 'Helvetica-Bold')
                elif status == "En congé":
                    table_style.add('BACKGROUND', (2, i), (2, i), colors.plum)
                    table_style.add('TEXTCOLOR', (2, i), (2, i), colors.purple)
                    table_style.add('FONTNAME', (2, i), (2, i), 'Helvetica-Bold')
                else:  # Non défini
                    table_style.add('BACKGROUND', (2, i), (2, i), colors.lightyellow)
                    table_style.add('TEXTCOLOR', (2, i), (2, i), colors.darkorange)
                    table_style.add('FONTNAME', (2, i), (2, i), 'Helvetica-Bold')

                # Highlight grand déplacement
                if data[i][4] == "Oui":
                    table_style.add('BACKGROUND', (4, i), (4, i), colors.lightblue)
                    table_style.add('TEXTCOLOR', (4, i), (4, i), colors.darkblue)
                    table_style.add('FONTNAME', (4, i), (4, i), 'Helvetica-Bold')

            table.setStyle(table_style)
            elements.append(table)

            elements.append(Spacer(1, 30))

            # Add summary statistics in two columns
            elements.append(Paragraph("Résumé de la journée", subtitle_style))
            elements.append(Spacer(1, 10))

            total_staff = staff_members.count()
            present_staff = sum(
                1 for s in staff_members if attendance_dict.get(s.id) and attendance_dict[s.id].present is True)
            absent_staff = sum(
                1 for s in staff_members if
                attendance_dict.get(s.id) and attendance_dict[s.id].present is False and attendance_dict[
                    s.id].absence_reason != 'CONGE_STATUS')
            conge_staff = sum(
                1 for s in staff_members if
                attendance_dict.get(s.id) and attendance_dict[s.id].absence_reason == 'CONGE_STATUS')
            undefined_staff = total_staff - present_staff - absent_staff - conge_staff

            # Calculate total hours worked and grand déplacements - FIX: Handle Decimal properly
            total_hours = 0
            for s in staff_members:
                attendance = attendance_dict.get(s.id)
                if attendance and attendance.present is True and attendance.hours_worked:
                    # Convert Decimal to float
                    total_hours += float(attendance.hours_worked)

            grand_deplacement_count = sum(
                1 for s in staff_members
                if attendance_dict.get(s.id) and attendance_dict[s.id].present is True and attendance_dict[
                    s.id].grand_deplacement
            )

            present_percentage = round(present_staff / total_staff * 100, 1) if total_staff > 0 else 0

            # Create summary in table format for better layout
            summary_data = [
                ["📊 Statistiques Générales", "Valeur", "📈 Détails Avancés", "Valeur"],
                ["Personnel total", f"{total_staff}", "Heures travaillées", f"{total_hours}h"],
                ["Personnel présent", f"{present_staff} ({present_percentage}%)", "Grands déplacements",
                 f"{grand_deplacement_count}"],
                ["Personnel absent", f"{absent_staff}", "Taux de présence", f"{present_percentage}%"],
                ["Personnel en congé", f"{conge_staff}", "Personnel non défini", f"{undefined_staff}"],
            ]

            summary_table = Table(summary_data, colWidths=[150, 80, 150, 80])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))

            elements.append(summary_table)
            elements.append(Spacer(1, 20))

            # Add city breakdown
            elements.append(Paragraph("Détail par chantier", subtitle_style))
            elements.append(Spacer(1, 10))

            cities = City.objects.all()
            city_data = [["Chantier", "Total", "Présents", "Absents", "Congés", "Heures", "Grands Dép."]]

            for city in cities:
                city_staff = staff_members.filter(city=city)
                city_total = city_staff.count()

                city_present = sum(
                    1 for s in city_staff if attendance_dict.get(s.id) and attendance_dict[s.id].present is True)
                city_absent = sum(
                    1 for s in city_staff if
                    attendance_dict.get(s.id) and attendance_dict[s.id].present is False and attendance_dict[
                        s.id].absence_reason != 'CONGE_STATUS')
                city_conge = sum(
                    1 for s in city_staff if
                    attendance_dict.get(s.id) and attendance_dict[s.id].absence_reason == 'CONGE_STATUS')

                # Calculate city hours and grand déplacements - FIX: Handle Decimal properly
                city_hours = 0
                for s in city_staff:
                    attendance = attendance_dict.get(s.id)
                    if attendance and attendance.present is True and attendance.hours_worked:
                        # Convert Decimal to float
                        city_hours += float(attendance.hours_worked)

                city_grand_deplacement = sum(
                    1 for s in city_staff
                    if attendance_dict.get(s.id) and attendance_dict[s.id].present is True and attendance_dict[
                        s.id].grand_deplacement
                )

                city_data.append([
                    city.name,
                    str(city_total),
                    str(city_present),
                    str(city_absent),
                    str(city_conge),
                    f"{city_hours}h" if city_hours > 0 else "0h",
                    str(city_grand_deplacement)
                ])

            city_table = Table(city_data, colWidths=[120, 50, 60, 60, 60, 60, 70])
            city_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))

            elements.append(city_table)

            # Build PDF document
            doc.build(elements)

            # Get the value of the BytesIO buffer
            pdf = buffer.getvalue()
            buffer.close()

            # Create the HttpResponse with PDF
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="rapport-presences-paysage.pdf"'
            response.write(pdf)

            return response

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


# Ajouter cette fonction à staff_attendance/views.py

# Ajouter cette fonction à staff_attendance/views.py

@csrf_exempt
@login_required
def clear_attendance(request):
    """API endpoint to clear staff attendance status"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            staff_id = data.get('staff_id')

            if not staff_id:
                return JsonResponse({'success': False, 'message': 'ID du personnel requis'}, status=400)

            staff = StaffMember.objects.get(id=staff_id)
            today = timezone.now().date()

            # Try to find and delete attendance record for today
            try:
                attendance = Attendance.objects.get(
                    staff_member=staff,
                    date=today
                )
                attendance.delete()

                return JsonResponse({
                    'success': True,
                    'message': f'Statut de {staff.name} effacé avec succès'
                })

            except Attendance.DoesNotExist:
                return JsonResponse({
                    'success': True,
                    'message': f'Aucun statut à effacer pour {staff.name}'
                })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
@csrf_exempt
def generate_history_pdf(request):
    """API endpoint to generate a PDF report of attendance history in landscape format"""
    if request.method == 'POST':
        try:
            # Get month and year from request
            data = json.loads(request.body) if request.body else {}
            month = data.get('month')
            year = data.get('year')

            # If no month/year provided, use current month
            if not month or not year:
                today = timezone.now().date()
                month = today.month
                year = today.year
            else:
                month = int(month)
                year = int(year)

            # Calculate the first and last day of the requested month
            import calendar
            start_date = datetime.date(year, month, 1)
            last_day = calendar.monthrange(year, month)[1]
            end_date = datetime.date(year, month, last_day)

            # Get all staff with their attendance
            staff_members = StaffMember.objects.select_related('city').all()

            # Get attendance records for the period
            attendances = Attendance.objects.filter(
                date__gte=start_date,
                date__lte=end_date
            ).order_by('-date')

            # Create a dictionary for quick lookup
            attendance_dict = {}
            for attendance in attendances:
                key = (attendance.date, attendance.staff_member_id)
                attendance_dict[key] = attendance

            # Create a BytesIO buffer for the PDF
            buffer = BytesIO()

            # Create the PDF object in LANDSCAPE mode
            from reportlab.lib.pagesizes import A4, landscape
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                topMargin=20,
                bottomMargin=20,
                leftMargin=20,
                rightMargin=20
            )
            elements = []

            # Set up styles
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']

            # Add title
            month_name = calendar.month_name[month]
            elements.append(Paragraph(f"Historique de Présence - {month_name} {year}", title_style))
            elements.append(Spacer(1, 12))

            # Add date range and summary in one line
            total_days = (end_date - start_date).days + 1
            working_days = 0
            current_date = start_date
            while current_date <= end_date:
                if current_date.weekday() < 5:  # Monday = 0, Sunday = 6
                    working_days += 1
                current_date += datetime.timedelta(days=1)

            date_and_summary = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')} | {total_days} jours total, {working_days} jours ouvrables"
            elements.append(Paragraph(date_and_summary, normal_style))
            elements.append(Spacer(1, 15))

            # Get all dates in the range
            dates = []
            current_date = start_date
            while current_date <= end_date:
                dates.append(current_date)
                current_date += datetime.timedelta(days=1)

            # Create single table with all staff (no grouping by city)
            # Create table header with dates
            header = ["Nom du personnel"]
            for date in dates:
                # Format: day + weekday initial
                day_str = date.strftime('%d')
                weekday_initial = date.strftime('%a')[0].upper()  # M, T, W, T, F, S, S

                # Special formatting for weekends
                if date.weekday() >= 5:  # Saturday or Sunday
                    header.append(f"{day_str}\n{weekday_initial}*")
                else:
                    header.append(f"{day_str}\n{weekday_initial}")

            # Create table data
            data = [header]

            # Add staff rows - ALL STAFF IN ONE TABLE
            for staff in staff_members:
                row = [staff.name]

                # Add status for each date
                for date in dates:
                    key = (date, staff.id)
                    attendance = attendance_dict.get(key)

                    if attendance:
                        if attendance.absence_reason == 'CONGE_STATUS':
                            status = "C"  # Congé
                        elif attendance.present is True:
                            # Show hours if available
                            if attendance.hours_worked:
                                hours_float = float(attendance.hours_worked)
                                if hours_float.is_integer():
                                    hours = int(hours_float)
                                else:
                                    hours = hours_float
                                status = f"P{hours}"
                                # Add G for grand déplacement
                                if attendance.grand_deplacement:
                                    status += "G"
                            else:
                                status = "P"
                                if attendance.grand_deplacement:
                                    status += "G"
                        elif attendance.present is False:
                            if attendance.absence_reason and len(attendance.absence_reason) > 0:
                                status = attendance.absence_reason  # Show exact reason: ABNJ, ABA, AM
                            else:
                                status = "A"  # Generic absent
                        else:
                            status = "-"  # Undefined
                    else:
                        status = "-"  # No record

                    row.append(status)

                data.append(row)

            # Calculate column widths for landscape mode
            available_width = landscape(A4)[0] - 40  # Subtract margins
            name_col_width = 120  # Fixed width for name column
            date_col_width = (available_width - name_col_width) / len(dates)
            date_col_width = max(date_col_width, 18)  # Minimum width

            col_widths = [name_col_width] + [date_col_width] * len(dates)

            # Create the table
            table = Table(data, colWidths=col_widths)

            # Style the table
            table_style = TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),

                # Body styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),

                # Name column styling
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('LEFTPADDING', (0, 1), (0, -1), 8),
            ])

            # Add conditional formatting and weekend highlighting
            for j in range(1, len(header)):
                date_index = j - 1
                if date_index < len(dates):
                    date = dates[date_index]

                    # Highlight weekends
                    if date.weekday() >= 5:  # Saturday or Sunday
                        table_style.add('BACKGROUND', (j, 0), (j, 0), colors.darkgrey)
                        table_style.add('BACKGROUND', (j, 1), (j, -1), colors.lightsteelblue)

            # Add status color coding
            for i in range(1, len(data)):
                for j in range(1, len(data[i])):
                    cell_value = data[i][j]
                    if cell_value.startswith("P"):
                        # Present - green
                        table_style.add('BACKGROUND', (j, i), (j, i), colors.lightgreen)
                        table_style.add('TEXTCOLOR', (j, i), (j, i), colors.darkgreen)
                        table_style.add('FONTNAME', (j, i), (j, i), 'Helvetica-Bold')
                    elif cell_value in ["A", "ABNJ", "ABA", "AM"] or cell_value.startswith("A"):
                        # Absent - red                        # Absent - red
                        table_style.add('BACKGROUND', (j, i), (j, i), colors.lightcoral)
                        table_style.add('TEXTCOLOR', (j, i), (j, i), colors.darkred)
                        table_style.add('FONTNAME', (j, i), (j, i), 'Helvetica-Bold')
                    elif cell_value == "C":
                        # Congé - purple
                        table_style.add('BACKGROUND', (j, i), (j, i), colors.plum)
                        table_style.add('TEXTCOLOR', (j, i), (j, i), colors.purple)
                        table_style.add('FONTNAME', (j, i), (j, i), 'Helvetica-Bold')

            table.setStyle(table_style)
            elements.append(table)
            elements.append(Spacer(1, 20))

            # Add legend
            elements.append(Paragraph("Légende des codes:", subtitle_style))
            legend_style = styles['Normal']
            legend_style.fontSize = 10

            elements.append(Paragraph(
                "• <b>P</b> = Présent | <b>P8</b> = Présent 8h | <b>P8G</b> = Présent 8h + Grand déplacement (nuit)",
                legend_style))
            elements.append(Paragraph("• <b>A</b> = Absent | <b>ABNJ/ABA/AM</b> = Motifs d'absence spécifiques", legend_style))
            elements.append(Paragraph("• <b>C</b> = En congé | <b>-</b> = Non défini (pas de données)", legend_style))
            elements.append(Paragraph("• <b>*</b> = Week-end (samedi/dimanche)", legend_style))
            elements.append(Spacer(1, 15))

            # Add statistics section
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("Statistiques de la période:", subtitle_style))

            # Calculate period statistics
            # Calculate period statistics - ONLY FOR THE SELECTED MONTH
            total_present_days = 0
            total_hours_worked = 0
            total_grand_deplacements = 0
            total_absences = 0
            total_conges = 0
            total_staff_count = len(staff_members)

            # Filter attendances to only the selected month/year
            month_attendances = attendances.filter(
                date__year=year,
                date__month=month
            )

            for attendance in month_attendances:  # ← CORRECTION ICI
                if attendance.present is True:
                    total_present_days += 1
                    if attendance.hours_worked:
                        # Convert Decimal to float
                        total_hours_worked += float(attendance.hours_worked)
                    if attendance.grand_deplacement:
                        total_grand_deplacements += 1
                elif attendance.present is False:
                    total_absences += 1
                elif attendance.absence_reason == 'CONGE_STATUS':
                    total_conges += 1

            # Calculate averages
            period_days = (end_date - start_date).days + 1
            avg_presence_rate = (total_present_days / (total_staff_count * period_days) * 100) if total_staff_count > 0 else 0
            avg_hours_per_day = (total_hours_worked / total_present_days) if total_present_days > 0 else 0

            stats_style = ParagraphStyle(
                'StatsStyle',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=6
            )

            elements.append(Paragraph(
                f"📊 <b>Période:</b> {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')} ({period_days} jours)",
                stats_style))
            elements.append(Paragraph(
                f"📊 <b>Jours de présence totaux:</b> {total_present_days} | <b>Heures travaillées totales:</b> {total_hours_worked:.1f}h",
                stats_style))
            elements.append(Paragraph(
                f"📊 <b>Grands déplacements:</b> {total_grand_deplacements} | <b>Absences:</b> {total_absences} | <b>Congés:</b> {total_conges}",
                stats_style))
            elements.append(Paragraph(
                f"📊 <b>Taux de présence moyen:</b> {avg_presence_rate:.1f}% | <b>Heures moyennes/jour:</b> {avg_hours_per_day:.1f}h",
                stats_style))

            elements.append(Spacer(1, 15))

            # Add statistics section
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("Statistiques de la période:", subtitle_style))

            # Filter to only the requested month/year
            month_attendances = [att for att in attendances if att.date.year == year and att.date.month == month]

            # Calculate statistics by chantier
            elements.append(Spacer(1, 15))
            elements.append(Paragraph("Statistiques par chantier:", subtitle_style))

            chantier_stats = {}
            for attendance in month_attendances:  # ← CORRECTION ICI
                if attendance.present is True and attendance.staff_member.city:
                    chantier_name = attendance.staff_member.city.name
                    hours_worked = float(attendance.hours_worked) if attendance.hours_worked else 0

                    if chantier_name not in chantier_stats:
                        chantier_stats[chantier_name] = {
                            'total_hours': 0,
                            'present_days': 0,
                            'staff_count': set(),
                            'grand_deplacements': 0
                        }

                    chantier_stats[chantier_name]['total_hours'] += hours_worked
                    chantier_stats[chantier_name]['present_days'] += 1
                    chantier_stats[chantier_name]['staff_count'].add(attendance.staff_member.id)

                    if attendance.grand_deplacement:
                        chantier_stats[chantier_name]['grand_deplacements'] += 1

            # Create chantier statistics table
            if chantier_stats:
                chantier_data = [["Chantier", "Heures totales", "Jours présence", "Personnel unique", "Grands dépl."]]

                # Sort by total hours (descending)
                sorted_chantiers = sorted(chantier_stats.items(), key=lambda x: x[1]['total_hours'], reverse=True)

                for chantier_name, stats in sorted_chantiers:
                    chantier_data.append([
                        chantier_name,
                        f"{stats['total_hours']:.1f}h",
                        str(stats['present_days']),
                        str(len(stats['staff_count'])),
                        str(stats['grand_deplacements'])
                    ])

                # Create the table
                chantier_table = Table(chantier_data, colWidths=[200, 80, 80, 80, 80])
                chantier_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Chantier names left-aligned
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))

                elements.append(chantier_table)
                elements.append(Spacer(1, 10))

                # Add summary
                total_chantiers = len(chantier_stats)
                most_active_chantier = max(chantier_stats.items(), key=lambda x: x[1]['total_hours'])

                elements.append(Paragraph(
                    f"📈 <b>Chantiers actifs:</b> {total_chantiers} | <b>Plus actif:</b> {most_active_chantier[0]} ({most_active_chantier[1]['total_hours']:.1f}h)",
                    stats_style))
            else:
                elements.append(Paragraph("Aucune donnée de chantier disponible pour cette période.", stats_style))

            elements.append(Spacer(1, 15))

            # Build PDF document
            doc.build(elements)

            # Get the value of the BytesIO buffer
            pdf = buffer.getvalue()
            buffer.close()

            # Create the HttpResponse with PDF
            response = HttpResponse(content_type='application/pdf')
            filename = f'historique-presences-{year}-{month:02d}-paysage.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response.write(pdf)

            return response

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
@login_required
def mark_absence_period(request):
    """API endpoint to mark a staff member as absent for a period"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            staff_id = data.get('staff_id')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            reason = data.get('reason', 'ABNJ')

            staff = StaffMember.objects.get(id=staff_id)

            # Parse dates
            start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()

            # Validate dates
            if start_date_obj > end_date_obj:
                return JsonResponse(
                    {'success': False, 'message': 'La date de début doit être antérieure à la date de fin'}, status=400)

            # Create attendance records for each day
            current_date = start_date_obj
            days_count = 0

            while current_date <= end_date_obj:
                attendance, created = Attendance.objects.get_or_create(
                    staff_member=staff,
                    date=current_date,
                    defaults={
                        'present': False,
                        'timestamp': timezone.now(),
                        'absence_reason': reason,
                        'created_by': request.user
                    }
                )

                if not created:
                    attendance.present = False
                    attendance.absence_reason = reason
                    attendance.timestamp = timezone.now()
                    attendance.updated_by = request.user
                    attendance.save()

                days_count += 1
                current_date += datetime.timedelta(days=1)

            return JsonResponse({
                'success': True,
                'message': f'{staff.name} marqué absent ({reason}) du {start_date_obj.strftime("%d/%m/%Y")} au {end_date_obj.strftime("%d/%m/%Y")} ({days_count} jours)'
            })

        except StaffMember.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Format de date invalide'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
# Solutions pour améliorer la lisibilité du PDF historique mensuel

@csrf_exempt
def generate_excel_history_protected(request):
    """Excel historique mensuel avec protection"""
    if request.method == 'POST':
        try:
            # Récupération des paramètres
            data = json.loads(request.body) if request.body else {}
            month = int(data.get('month', timezone.now().date().month))
            year = int(data.get('year', timezone.now().date().year))

            # Calcul des dates
            start_date = datetime.date(year, month, 1)
            last_day = calendar.monthrange(year, month)[1]
            end_date = datetime.date(year, month, last_day)

            # Récupération des données
            staff_members = StaffMember.objects.select_related('city').all()
            attendances = Attendance.objects.filter(
                date__gte=start_date, date__lte=end_date
            ).order_by('staff_member__name', 'date')

            # Organisation des données
            staff_data = {}
            for staff in staff_members:
                staff_data[staff.id] = {'staff': staff, 'attendances': {}}
            for attendance in attendances:
                staff_data[attendance.staff_member_id]['attendances'][attendance.date] = attendance

            # Création du workbook
            wb = openpyxl.Workbook()
            ws = wb.active

            mois_fr = MOIS_FRANCAIS.get(month, str(month))
            ws.title = f"{mois_fr} {year}"

            # Fonction pour convertir un numéro de colonne en lettre Excel
            def get_excel_column(col_num):
                """Convertit un numéro de colonne en lettre Excel (1->A, 27->AA, etc.)"""
                result = ""
                while col_num > 0:
                    col_num -= 1
                    result = chr(65 + (col_num % 26)) + result
                    col_num //= 26
                return result

            # Calculer la dernière colonne
            total_columns = 2 + last_day + 3  # Personnel + Chantier + jours + 3 totaux
            last_column = get_excel_column(total_columns)

            # Styles avec polices plus grandes pour Excel
            header_font = Font(bold=True, color="FFFFFF", size=12)  # Plus grand
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            present_font = Font(color="006400", bold=True, size=11)  # Vert foncé, plus grand
            absent_font = Font(color="DC143C", bold=True, size=10)  # Rouge foncé, plus grand
            conge_font = Font(color="808080", size=10)  # Gris, plus grand
            grand_dep_font = Font(color="0000FF", bold=True, size=11)  # Bleu, plus grand
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Titre - Plus grand et plus visible
            ws.merge_cells(f'A1:{last_column}1')
            title_cell = ws.cell(row=1, column=1, value=f"Historique des Présences - {mois_fr} {year}")
            title_cell.font = Font(bold=True, size=18, color="1F4E79")  # Plus grand et coloré
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajuster la hauteur de la ligne du titre
            ws.row_dimensions[1].height = 25

            # En-têtes avec plus d'espace
            headers = ["Personnel", "Chantier"]
            for day in range(1, last_day + 1):
                headers.append(str(day))
            headers.extend(["Total Présences", "Total Absences" ])

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Ajuster la hauteur de la ligne d'en-tête
            ws.row_dimensions[3].height = 20

            # Données par personnel
            row = 4
            for staff_id, staff_info in staff_data.items():
                staff = staff_info['staff']
                attendances_dict = staff_info['attendances']

                # Nom du personnel - CASE BLEUE pour grand déplacement
                has_grand_deplacement = any(att.grand_deplacement for att in attendances_dict.values() if att.present)
                name_cell = ws.cell(row=row, column=1, value=staff.name)
                name_cell.font = Font(size=11)  # Nom normal
                if has_grand_deplacement:
                    # CASE BLEUE pour grand déplacement
                    name_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                name_cell.border = border
                name_cell.alignment = Alignment(vertical="center")

                # Chantier avec police plus grande
                chantier = staff.city.name if staff.city else "N/A"
                chantier_cell = ws.cell(row=row, column=2, value=chantier)
                chantier_cell.font = Font(size=10)  # Police plus grande
                chantier_cell.border = border
                chantier_cell.alignment = Alignment(vertical="center")

                # Statuts par jour
                present_count = absent_count  = 0

                for day in range(1, last_day + 1):
                    current_date = datetime.date(year, month, day)
                    attendance = attendances_dict.get(current_date)
                    col = day + 2  # Décalage pour Personnel et Chantier

                    if attendance:
                        if attendance.present:
                            cell = ws.cell(row=row, column=col, value="PRÉSENT")
                            # CASE VERTE avec texte blanc pour présent
                            cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # Vert
                            cell.font = Font(color="FFFFFF", bold=True, size=8)  # Blanc
                            present_count += 1

                        elif attendance.absence_reason == 'CONGE_STATUS':
                            cell = ws.cell(row=row, column=col, value="CONGÉ")
                            # CASE VIOLETTE avec texte blanc pour congé
                            cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD",
                                                    fill_type="solid")  # Violet
                            cell.font = Font(color="FFFFFF", bold=True, size=8)  # Blanc
                        elif attendance.present is False and attendance.absence_reason in ['ABNJ', 'AM', 'ABA']:
                            # AFFICHER LE TYPE D'ABSENCE EXACT avec case rouge et texte blanc
                            cell = ws.cell(row=row, column=col, value=attendance.absence_reason)
                            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C",
                                                    fill_type="solid")  # Rouge
                            cell.font = Font(color="FFFFFF", bold=True, size=8)  # Blanc
                            absent_count += 1
                        elif attendance.present is False:
                            # Si absent mais pas de raison spécifique - case rouge
                            cell = ws.cell(row=row, column=col, value="ABSENT")
                            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C",
                                                    fill_type="solid")  # Rouge
                            cell.font = Font(color="FFFFFF", bold=True, size=8)  # Blanc
                            absent_count += 1
                        else:
                            cell = ws.cell(row=row, column=col, value="-")
                            cell.font = Font(size=8, color="666666")
                    else:
                        cell = ws.cell(row=row, column=col, value="-")
                        cell.font = Font(size=8, color="666666")

                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                # Totaux avec police plus grande
                total_p_cell = ws.cell(row=row, column=last_day + 3, value=present_count)
                total_p_cell.alignment = Alignment(horizontal="center", vertical="center")
                total_p_cell.border = border
                total_p_cell.font = Font(bold=True, size=10)

                total_a_cell = ws.cell(row=row, column=last_day + 4, value=absent_count)
                total_a_cell.alignment = Alignment(horizontal="center", vertical="center")
                total_a_cell.border = border
                total_a_cell.font = Font(bold=True, size=10)



                row += 1

            # Ajustement des colonnes - ADAPTATION AUTOMATIQUE À LA LONGUEUR
            ws.column_dimensions['A'].width = 25  # Personnel
            ws.column_dimensions['B'].width = 20  # Chantier

            # Jours du mois - Largeur adaptative selon le contenu
            for day in range(1, last_day + 1):
                col_letter = get_excel_column(day + 2)
                # Largeur pour accommoder "PRÉSENT" (8 chars) et "ABNJ" (4 chars)
                ws.column_dimensions[col_letter].width = 8

            # Colonnes totaux
            for i in range(3):
                col_letter = get_excel_column(last_day + 3 + i)
                ws.column_dimensions[col_letter].width = 10

            # AJUSTEMENT AUTOMATIQUE - Parcourir toutes les cellules pour adapter
            for col in range(3, last_day + 3):  # Colonnes des jours
                max_length = 0
                col_letter = get_excel_column(col)

                for row_num in range(3, row):  # De la ligne d'en-tête aux données
                    cell_value = ws.cell(row=row_num, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                # Ajuster la largeur avec une marge
                ws.column_dimensions[col_letter].width = max(max_length + 2, 6)

            # PROTECTION COMPLÈTE
            ws.protection = SheetProtection(
                password='admin123',
                sheet=True,
                objects=True,
                scenarios=True,
                formatCells=False,
                formatColumns=False,
                formatRows=False,
                insertColumns=False,
                insertRows=False,
                deleteColumns=False,
                deleteRows=False,
                selectLockedCells=True,
                sort=False,
                autoFilter=False
            )

            wb.security = WorkbookProtection(
                workbookPassword='admin123',
                lockStructure=True,
                lockWindows=True
            )

            # Légende mise à jour avec les nouvelles couleurs
            legend_row = row + 1
            ws.merge_cells(f'A{legend_row}:{last_column}{legend_row}')
            legend_cell = ws.cell(row=legend_row, column=1,
                                  value="🟢 VERT = Présent | 🔴 ROUGE = Absent/ABNJ/AM/ABA | 🟣 VIOLET = Congé | - = Pas de données | 🔵 CASE BLEUE = Grands déplacements")
            legend_cell.font = Font(italic=True, size=11, color="555555")
            legend_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[legend_row].height = 20

            # Sauvegarde
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'historique-{mois_fr.lower()}-{year}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)



# UTILISATION: Dans la fonction principale, décommentez l'option que vous préférez :
# Option 1: Ultra-compact mais tout sur une page
# Option 2: Par semaines (plus lisible, plusieurs pages)
# Option 3: Résumé avec statistiques (le plus clair)



@csrf_exempt
def get_cities_list(request):
    """API endpoint to get the list of all cities with their IDs"""
    try:
        cities = City.objects.all().order_by('name')
        cities_data = []
        for city in cities:
            staff_count = StaffMember.objects.filter(city=city).count()
            cities_data.append({
                'id': city.id,
                'name': city.name,
                'staff_count': staff_count
            })

        return JsonResponse({'success': True, 'cities': cities_data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def add_city(request):
    """API endpoint to add a new city to a zone"""
    if request.method == 'POST':
        data = json.loads(request.body)
        city_name = data.get('name', '').strip()
        zone_id = data.get('zone_id')

        if not city_name:
            return JsonResponse({'success': False, 'message': 'Le nom du chantier est requis'}, status=400)

        if not zone_id:
            return JsonResponse({'success': False, 'message': 'La zone est requise'}, status=400)

        # Check if city already exists
        if City.objects.filter(name__iexact=city_name).exists():
            return JsonResponse({'success': False, 'message': 'Ce chantier existe déjà'}, status=400)

        try:
            zone = Zone.objects.get(id=zone_id)
            city = City.objects.create(name=city_name, zone=zone)
            return JsonResponse({
                'success': True,
                'message': f'Chantier "{city_name}" ajouté à la zone "{zone.name}" avec succès',
                'city': {
                    'id': city.id,
                    'name': city.name,
                    'zone_id': zone.id,
                    'zone_name': zone.name,
                    'staff_count': 0
                }
            })
        except Zone.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Zone non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)




@csrf_exempt
def update_city_name(request):
    """API endpoint to update a city's name"""
    if request.method == 'POST':
        data = json.loads(request.body)
        city_id = data.get('city_id')
        new_name = data.get('new_name', '').strip()

        if not new_name:
            return JsonResponse({'success': False, 'message': 'Le nouveau nom est requis'}, status=400)

        if not city_id:
            return JsonResponse({'success': False, 'message': 'ID de la ville requis'}, status=400)

        try:
            city = City.objects.get(id=city_id)

            # Check if new name already exists (excluding current city)
            if City.objects.filter(name__iexact=new_name).exclude(id=city_id).exists():
                return JsonResponse({'success': False, 'message': 'Une ville avec ce nom existe déjà'}, status=400)

            old_name = city.name
            city.name = new_name
            city.save()

            return JsonResponse({
                'success': True,
                'message': f'Ville renommée de "{old_name}" à "{new_name}"',
                'city': {
                    'id': city.id,
                    'name': city.name,
                    'old_name': old_name
                }
            })
        except City.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Ville non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def delete_city(request):
    """API endpoint to delete a city"""
    if request.method == 'POST':
        data = json.loads(request.body)
        city_id = data.get('city_id')

        if not city_id:
            return JsonResponse({'success': False, 'message': 'ID de la ville requis'}, status=400)

        try:
            city = City.objects.get(id=city_id)

            # Check if city has staff members
            staff_count = StaffMember.objects.filter(city=city).count()
            if staff_count > 0:
                return JsonResponse({
                    'success': False,
                    'message': f'Impossible de supprimer "{city.name}". Elle contient {staff_count} membre(s) du personnel.'
                }, status=400)

            city_name = city.name
            city.delete()

            return JsonResponse({
                'success': True,
                'message': f'Ville "{city_name}" supprimée avec succès',
                'deleted_city_name': city_name
            })
        except City.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Ville non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)


@csrf_exempt
def get_zones_list(request):
    """API endpoint to get the list of all zones with their cities"""
    try:
        zones = Zone.objects.prefetch_related('cities').all().order_by('name')
        zones_data = []

        for zone in zones:
            cities_in_zone = []
            for city in zone.cities.all():
                staff_count = StaffMember.objects.filter(city=city).count()
                cities_in_zone.append({
                    'id': city.id,
                    'name': city.name,
                    'staff_count': staff_count
                })

            zones_data.append({
                'id': zone.id,
                'name': zone.name,
                'cities': cities_in_zone,
                'cities_count': len(cities_in_zone)
            })

        return JsonResponse({'success': True, 'zones': zones_data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


